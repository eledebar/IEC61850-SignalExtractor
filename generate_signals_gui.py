import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess
import sys

NS = {'scl': 'http://www.iec.ch/61850/2003/SCL'}

def ensure_dependencies():
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

def build_signal_name(fcda):
    prefix = fcda.get('prefix', '')
    ln_class = fcda.get('lnClass', '')
    ln_inst = fcda.get('lnInst', '')
    do_name = fcda.get('doName', '')
    da_name = fcda.get('daName')
    fc = fcda.get('fc', '')

    name = f"{prefix}{ln_class}{ln_inst}.{do_name}"
    if da_name:
        name += f".{da_name}"

    if fc == "MX":
        name += ".mag.f"
    elif fc in ["ST", "SP"]:
        name += ".stVal"
    elif fc == "CO":
        name += ".ctlVal"

    return name

def extract_datasets(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    datasets = defaultdict(list)

    for dataset in root.findall('.//scl:DataSet', NS):
        name = dataset.get('name', 'Unnamed')
        for fcda in dataset.findall('scl:FCDA', NS):
            signal = build_signal_name(fcda)
            datasets[name].append(signal)

    return datasets

def export_to_excel(columns, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Signals"

    header_font = Font(bold=True, name='Calibri')
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = list(columns.keys())
    ws.append(headers)

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    max_rows = max(len(lst) for lst in columns.values())
    for row in range(max_rows):
        for col_idx, key in enumerate(headers, 1):
            value = columns[key][row] if row < len(columns[key]) else ''
            cell = ws.cell(row=row + 2, column=col_idx, value=value)
            cell.border = border

    for col_idx in range(1, len(headers) + 1):
        length = max(len(str(cell.value or '')) for cell in ws[get_column_letter(col_idx)])
        ws.column_dimensions[get_column_letter(col_idx)].width = length + 2

    ws.freeze_panes = "A2"
    wb.save(filename)

def main():
    ensure_dependencies()

    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("IEC 61850 Signal Extractor", "Please select an IEC 61850 XML file")

    file_path = filedialog.askopenfilename(title="Choose SCL XML File", filetypes=[("XML files", "*.xml")])

    if not file_path:
        messagebox.showwarning("No file", "No file selected. Exiting.")
        return

    try:
        data = extract_datasets(file_path)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"signals_{timestamp}.xlsx"
        export_to_excel(data, output_file)
        messagebox.showinfo("Done", f"Signals exported to:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    main()
